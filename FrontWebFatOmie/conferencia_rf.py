"""
Conferência de RFs (fluxo VALE) — Front Streamlit (Sillion)

Responsabilidades:
1. Extrair a coluna "RF" do arquivo enviado (xlsx / xlsb / csv, cabeçalho na linha 1).
2. Consultar o webhook N8N de conferência com a lista de RFs e receber os já
   cadastrados no banco de dados (com a referência de onde estão no banco).
3. Gerar, quando necessário, um arquivo filtrado SEM as linhas dos RFs que o
   usuário optou por não reenviar — preservando o mesmo formato/colunas do
   arquivo original, para não quebrar a esteira existente no N8N.

Contrato do webhook de consulta (N8N):
  Request  (POST JSON): {"tipo_faturamento": "VALE", "rfs": ["6203134747", ...]}
  Response (JSON) — qualquer um dos formatos abaixo é aceito:
    a) {"existentes": [{"rf": "6203134747", "linha": 123}, ...]}
    b) [{"rf": "6203134747", "linha": 123}, ...]
    c) ["6203134747", ...]                       (sem a referência de linha)
  Chaves aceitas para o RF:    rf, RF, numero_rf
  Chaves aceitas para a linha: linha, linha_banco, row, registro, id
"""

import io
from datetime import datetime

import requests

RF_COLUNA = "RF"
TIMEOUT_CONSULTA = 60  # segundos

# Status de nota devolvidos pelo backend. Medido na Omie em 10/08/2026: o
# cStatusNFSe tem só dois valores, F (faturada) e C (cancelada) -- não existe
# código para "substituída", que se detecta pelo mesmo RF em 2 notas.
STATUS_FATURADA = "FATURADA"
STATUS_CANCELADA = "CANCELADA"
STATUS_PADRAO = STATUS_FATURADA  # quando o backend não informa


class ArquivoMemoria:
    """Imita a interface do UploadedFile do Streamlit (name + getvalue)."""

    def __init__(self, name: str, conteudo: bytes):
        self.name = name
        self._conteudo = conteudo

    def getvalue(self) -> bytes:
        return self._conteudo


# ============================================================
# Normalização
# ============================================================
def normalizar_rf(valor) -> str:
    """
    Converte o RF para string canônica para comparação:
    remove espaços e sufixo '.0' de floats vindos do Excel.
    Retorna "" para valores vazios/None.
    """
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    texto = str(valor).strip()
    if texto.endswith(".0"):
        try:
            texto = str(int(float(texto)))
        except ValueError:
            pass
    return texto


def _ext(nome: str) -> str:
    return nome.rsplit(".", 1)[-1].lower() if "." in nome else ""


# ============================================================
# 1) Extração dos RFs
# ============================================================
def extrair_rfs(nome_arquivo: str, conteudo: bytes) -> list:
    """
    Lê o arquivo e retorna [{"linha": <nº da linha no arquivo>, "rf": <str>}].
    A linha é a linha real da planilha (cabeçalho = 1, dados a partir da 2).
    Linhas com RF vazio são ignoradas na conferência (mas permanecem no arquivo).
    Levanta ValueError se a coluna "RF" não existir.
    """
    ext = _ext(nome_arquivo)
    if ext == "xlsx":
        return _extrair_xlsx(conteudo)
    if ext == "xlsb":
        return _extrair_xlsb(conteudo)
    if ext == "csv":
        return _extrair_csv(conteudo)
    raise ValueError(f"Extensão não suportada para conferência: .{ext}")


def _indice_coluna_rf(cabecalho) -> int:
    for i, nome in enumerate(cabecalho):
        if nome is not None and str(nome).strip().upper() == RF_COLUNA:
            return i
    raise ValueError(
        f'Coluna "{RF_COLUNA}" não encontrada no cabeçalho (linha 1) do arquivo.'
    )


def _extrair_xlsx(conteudo: bytes) -> list:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    linhas = ws.iter_rows(values_only=True)
    cabecalho = next(linhas, None)
    if cabecalho is None:
        raise ValueError("Arquivo vazio.")
    idx = _indice_coluna_rf(cabecalho)

    resultado = []
    for n, row in enumerate(linhas, start=2):
        rf = normalizar_rf(row[idx] if idx < len(row) else None)
        if rf:
            resultado.append({"linha": n, "rf": rf})
    wb.close()
    return resultado


def _extrair_xlsb(conteudo: bytes) -> list:
    from pyxlsb import open_workbook

    with open_workbook(io.BytesIO(conteudo)) as wb:
        with wb.get_sheet(1) as ws:
            resultado = []
            idx = None
            for row in ws.rows():
                valores = [c.v for c in row]
                num_linha = row[0].r + 1 if row else 0  # pyxlsb: r é 0-based
                if idx is None:
                    idx = _indice_coluna_rf(valores)
                    continue
                rf = normalizar_rf(valores[idx] if idx < len(valores) else None)
                if rf:
                    resultado.append({"linha": num_linha, "rf": rf})
    return resultado


def _decodificar_csv(conteudo: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return conteudo.decode(enc)
        except UnicodeDecodeError:
            continue
    raise ValueError("Não foi possível decodificar o CSV.")


def _detectar_separador(linha_cabecalho: str) -> str:
    candidatos = [";", ",", "\t", "|"]
    return max(candidatos, key=lambda s: linha_cabecalho.count(s))


def _extrair_csv(conteudo: bytes) -> list:
    import csv as csv_mod

    texto = _decodificar_csv(conteudo)
    primeira_linha = texto.splitlines()[0] if texto.splitlines() else ""
    sep = _detectar_separador(primeira_linha)

    leitor = csv_mod.reader(io.StringIO(texto), delimiter=sep)
    cabecalho = next(leitor, None)
    if cabecalho is None:
        raise ValueError("Arquivo vazio.")
    idx = _indice_coluna_rf(cabecalho)

    resultado = []
    for n, row in enumerate(leitor, start=2):
        rf = normalizar_rf(row[idx] if idx < len(row) else None)
        if rf:
            resultado.append({"linha": n, "rf": rf})
    return resultado


# ============================================================
# 2) Consulta ao webhook N8N
# ============================================================
def consultar_rfs(url: str, rfs: list, tipo_faturamento: str = "VALE") -> dict:
    """
    Envia a lista de RFs ao N8N e retorna {rf_normalizado: info}, onde info é:

        {
          "bloqueado":    bool,        # tem nota vinculada -> NÃO pode ser enviado
          "legado":       bool,        # já foi enviado antes, mas sem nota vinculada
          "enviado_em":   str | None,  # quando foi enviado (caso legado)
          "notas":        [ {"numero", "status", "os", "emp"} ],
          "substituicao": bool,        # o mesmo RF em mais de uma nota
          "ref":          str,         # referência legível (campo "linha" do webhook)
        }

    O mesmo RF pode voltar em VÁRIAS linhas da resposta: é assim que a
    substituição se manifesta (nota cancelada + nota nova para o mesmo RF).
    Por isso as notas são acumuladas em lista, não sobrescritas.

    Tolera a resposta ANTIGA (`{"rf", "linha"}`, sem o campo `bloqueado`):
    nesse caso o RF cai como *legado* e o comportamento fica idêntico ao de
    antes — aviso com opção de reenvio, sem bloqueio. Assim o front continua
    funcionando se o webhook v2 sair do ar.

    Levanta requests.RequestException em falha de rede e ValueError em
    resposta inesperada.
    """
    resp = requests.post(
        url,
        json={"tipo_faturamento": tipo_faturamento, "rfs": rfs},
        timeout=TIMEOUT_CONSULTA,
        headers={"Content-Type": "application/json"},
    )
    resp.raise_for_status()

    try:
        dados = resp.json()
    except ValueError:
        raise ValueError(f"Resposta do webhook não é JSON: {resp.text[:300]}")

    if isinstance(dados, dict):
        dados = dados.get("existentes", dados.get("data", []))
    if dados is None:
        dados = []
    if not isinstance(dados, list):
        raise ValueError(f"Formato de resposta inesperado do webhook: {type(dados)}")

    existentes = {}
    for item in dados:
        if not isinstance(item, dict):
            # resposta mais simples possível: só a lista de RFs
            rf = normalizar_rf(item)
            if rf:
                existentes.setdefault(rf, _info_vazia())["legado"] = True
            continue

        rf = normalizar_rf(item.get("rf", item.get("RF", item.get("numero_rf"))))
        if not rf:
            continue

        ref = item.get(
            "linha",
            item.get("linha_banco", item.get("row", item.get("registro", item.get("id", "-")))),
        )

        info = existentes.setdefault(rf, _info_vazia())
        if ref not in (None, "") and info["ref"] == "-":
            info["ref"] = str(ref)

        bloqueado = bool(item.get("bloqueado", False))
        if bloqueado:
            info["bloqueado"] = True
            info["notas"].append({
                "numero": _texto(item.get("numero_nota")),
                "status": (_texto(item.get("status_nota")) or STATUS_PADRAO).upper(),
                "os": _texto(item.get("os")),
                "emp": _texto(item.get("emp")),
            })
        else:
            # sem "bloqueado" (webhook antigo) ou bloqueado=false -> legado
            info["legado"] = True
            if item.get("enviado_em"):
                info["enviado_em"] = _texto(item.get("enviado_em"))

        if item.get("substituicao"):
            info["substituicao"] = True

    # o mesmo RF em 2+ notas é, por si, a assinatura de uma substituição
    for info in existentes.values():
        if len(info["notas"]) > 1:
            info["substituicao"] = True
        # bloqueio vence legado: se há nota, o RF não passa
        if info["bloqueado"]:
            info["legado"] = False

    return existentes


def _info_vazia() -> dict:
    return {
        "bloqueado": False,
        "legado": False,
        "enviado_em": None,
        "notas": [],
        "substituicao": False,
        "ref": "-",
    }


def _texto(valor) -> str:
    return "" if valor is None else str(valor).strip()


def tem_cancelada(info: dict) -> bool:
    """True se alguma das notas do RF está cancelada."""
    return any(n.get("status") == STATUS_CANCELADA for n in info.get("notas", []))


def todas_canceladas(info: dict) -> bool:
    """
    True se o RF tem nota(s) e TODAS estão canceladas.
    Diferente de tem_cancelada: numa substituição (cancelada + faturada nova),
    tem_cancelada é True mas todas_canceladas é False — e aí o RF continua
    bloqueado, porque a nota substituta está faturada e procede.
    """
    notas = info.get("notas", [])
    return bool(notas) and all(n.get("status") == STATUS_CANCELADA for n in notas)


def motivo_bloqueio(info: dict) -> str:
    """Frase curta explicando ao usuário por que o RF não pode ser enviado."""
    if info.get("substituicao"):
        return "RF já usado em mais de uma nota (substituição)"
    if tem_cancelada(info):
        return "Nota cancelada — refaturar exige liberação"
    return "Já existe nota fiscal para este RF"


# ============================================================
# 3) Arquivo filtrado (remove linhas de RFs não selecionados)
# ============================================================
def gerar_arquivo_filtrado(nome_arquivo: str, conteudo: bytes, linhas_remover: set):
    """
    Retorna um ArquivoMemoria sem as linhas indicadas (números de linha reais
    da planilha, cabeçalho = 1). Preserva o formato original:
      - xlsx → xlsx (openpyxl.delete_rows, mantém estilos/colunas)
      - csv  → csv  (remove as linhas do texto original)
      - xlsb → convertido para xlsx (xlsb não pode ser gravado; o conteúdo e
               as colunas são preservados)
    Se não houver linhas a remover, retorna o arquivo original intacto.
    """
    if not linhas_remover:
        return ArquivoMemoria(nome_arquivo, conteudo)

    ext = _ext(nome_arquivo)
    if ext == "xlsx":
        return ArquivoMemoria(nome_arquivo, _filtrar_xlsx(conteudo, linhas_remover))
    if ext == "csv":
        return ArquivoMemoria(nome_arquivo, _filtrar_csv(conteudo, linhas_remover))
    if ext == "xlsb":
        novo_nome = nome_arquivo.rsplit(".", 1)[0] + ".xlsx"
        return ArquivoMemoria(novo_nome, _xlsb_para_xlsx_filtrado(conteudo, linhas_remover))
    raise ValueError(f"Extensão não suportada para filtro: .{ext}")


def _filtrar_xlsx(conteudo: bytes, linhas_remover: set) -> bytes:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb[wb.sheetnames[0]]
    # Remove de baixo para cima para não deslocar os índices
    for n in sorted(linhas_remover, reverse=True):
        ws.delete_rows(n)
    saida = io.BytesIO()
    wb.save(saida)
    return saida.getvalue()


def _filtrar_csv(conteudo: bytes, linhas_remover: set) -> bytes:
    texto = _decodificar_csv(conteudo)
    fim_linha = "\r\n" if "\r\n" in texto else "\n"
    linhas = texto.splitlines()
    mantidas = [l for i, l in enumerate(linhas, start=1) if i not in linhas_remover]
    return fim_linha.join(mantidas).encode("utf-8")


def _xlsb_para_xlsx_filtrado(conteudo: bytes, linhas_remover: set) -> bytes:
    import openpyxl
    from pyxlsb import open_workbook, convert_date

    novo = openpyxl.Workbook()
    ws_novo = novo.active
    with open_workbook(io.BytesIO(conteudo)) as wb:
        ws_novo.title = wb.sheets[0]
        with wb.get_sheet(1) as ws:
            for row in ws.rows():
                num_linha = row[0].r + 1 if row else 0
                if num_linha in linhas_remover:
                    continue
                ws_novo.append([c.v for c in row])
    saida = io.BytesIO()
    novo.save(saida)
    return saida.getvalue()


# ============================================================
# Orquestração (usada pelo app.py)
# ============================================================
def conferir(nome_arquivo: str, conteudo: bytes, url_consulta: str,
             tipo_faturamento: str = "VALE",
             liberar_canceladas: bool = False) -> dict:
    """
    Executa extração + consulta e separa os RFs do arquivo em baldes:

    {
      "rfs":        [{"linha": n, "rf": str}],              # todos os RFs do arquivo
      "bloqueados": [{"linha": n, "rf": str, "info": {...}}],  # tem nota -> NÃO envia
      "canceladas": [{"linha": n, "rf": str, "info": {...}}],  # ver liberar_canceladas
      "legado":     [{"linha": n, "rf": str, "info": {...}}],  # enviado antes, sem nota
      "novos":      [{"linha": n, "rf": str}],                 # nunca vistos -> envia
      "duplicados_no_arquivo": {rf: [linhas]},   # mesmo RF em 2+ linhas do arquivo
      "consultado_em": datetime,

      # compatibilidade: quem ainda espera o formato antigo
      "duplicados": bloqueados + canceladas + legado,
    }

    liberar_canceladas — TEMPORÁRIO (12/08/2026, vigente até o término da
    reforma): quando True, RF cujas notas estão TODAS canceladas sai do balde
    de bloqueados e vai para "canceladas" — o front oferece reenvio mediante
    marcação explícita. Nota FATURADA continua bloqueando sempre; substituição
    (cancelada + faturada) também, porque a substituta procede. Com False,
    comportamento original: cancelada bloqueia.

    Um RF pode aparecer em mais de uma LINHA do arquivo; cada linha entra no
    balde correspondente, porque o filtro do arquivo trabalha por linha.
    """
    rfs = extrair_rfs(nome_arquivo, conteudo)
    existentes = consultar_rfs(url_consulta, [r["rf"] for r in rfs], tipo_faturamento)

    bloqueados, canceladas, legado, novos = [], [], [], []
    for r in rfs:
        info = existentes.get(r["rf"])
        if info is None:
            novos.append(r)
        elif info["bloqueado"]:
            if liberar_canceladas and todas_canceladas(info):
                canceladas.append({**r, "info": info})
            else:
                bloqueados.append({**r, "info": info})
        else:
            legado.append({**r, "info": info})

    # mesmo RF em mais de uma linha do arquivo = risco de faturar em duplicidade
    linhas_por_rf = {}
    for r in rfs:
        linhas_por_rf.setdefault(r["rf"], []).append(r["linha"])
    duplicados_no_arquivo = {
        rf: linhas for rf, linhas in linhas_por_rf.items() if len(linhas) > 1
    }

    return {
        "rfs": rfs,
        "bloqueados": bloqueados,
        "canceladas": canceladas,
        "legado": legado,
        "novos": novos,
        "duplicados_no_arquivo": duplicados_no_arquivo,
        "consultado_em": datetime.now(),
        "duplicados": bloqueados + canceladas + legado,
    }


def linhas_bloqueadas(resultado: dict) -> set:
    """
    Números de linha do arquivo que precisam sair do envio OBRIGATORIAMENTE.
    O usuário não escolhe: RF com nota vinculada não vai, e ponto.
    """
    return {b["linha"] for b in resultado.get("bloqueados", [])}
